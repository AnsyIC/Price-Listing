# pricing_catalog_builder.py
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

XLSX_PATH = "../workflow_data/KWT报价单.xlsx"
OUT_JSON = "../workflow_data/pricing_catalog.json"
OUT_TXT = "../workflow_data/pricing_catalog.txt"

# 允许/规范化的中文维度（你可按需继续扩展）
DIM_NORMALIZE = {
    "反应": "次",
    "样本": "样本",
    "样品": "样本",
    "样": "样本",
    "点": "部位",
    "点位": "部位",
    "膜": "张",
    "项": "个",
    "片": "张",
}

CNY_RE = re.compile(r"([+−\-]?\s*\d+(?:\.\d+)?)\s*元")


def clean_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_dim(d: str) -> str:
    d = clean_str(d)
    d = re.sub(r"[()（）].*", "", d).strip()
    return DIM_NORMALIZE.get(d, d)


def parse_price_cell(price_cell: Any) -> Dict[str, Any]:
    """
    Parse:
      - '8元/笼/天（...）' -> unit_price=8, dims=['笼','天']
      - '+4元/笼/天' -> is_addon=True
      - '按xx报价/需询价' -> priced=False
    """
    raw = clean_str(price_cell)
    if not raw:
        return {"raw": raw, "priced": False, "unit_price": None, "unit_text": "", "dims": [], "is_addon": False}

    raw_norm = raw.replace("−", "-")
    m = CNY_RE.search(raw_norm)
    if not m:
        return {"raw": raw, "priced": False, "unit_price": None, "unit_text": "", "dims": [], "is_addon": False}

    unit_price = float(m.group(1).replace(" ", ""))
    is_addon = raw_norm.strip().startswith("+") or ("基础上" in raw_norm and "+" in raw_norm)

    after = raw_norm[m.end():]  # part after '元'
    after_no_paren = re.split(r"[（(]", after, maxsplit=1)[0]
    unit_text = ("元" + after_no_paren).strip()

    dims: List[str] = []
    if "/" in after_no_paren:
        parts = [p.strip() for p in after_no_paren.split("/") if p.strip()]
        dims = parts

    return {
        "raw": raw,
        "priced": True,
        "unit_price": unit_price,
        "unit_text": unit_text,
        "dims": dims,
        "is_addon": is_addon,
    }


def parse_unit_text(unit_cell: Any) -> Dict[str, Any]:
    """
    病理 sheet 常见：
      - '元/张' / '元/样'
      - 或直接写 '张' / '样'
    """
    raw = clean_str(unit_cell)
    if not raw:
        return {"unit_text": "", "dims": []}

    txt = re.split(r"[（(]", raw, maxsplit=1)[0].strip()
    if "/" in txt:
        parts = [p.strip() for p in txt.split("/") if p.strip()]
        # drop leading 元 if present
        if parts and (parts[0] == "元" or parts[0].endswith("元")):
            dims = parts[1:]
        else:
            dims = parts
        return {"unit_text": txt, "dims": dims}

    # bare unit like '样'/'张'
    return {"unit_text": f"元/{txt}" if not txt.startswith("元") else txt, "dims": [txt]}


def make_item(
    sheet: str,
    service: str,
    unit_price: Optional[float],
    unit_dims_raw: List[str],
    unit_text: str,
    notes: str = "",
    category1: str = "",
    category2: str = "",
    priced: bool = True,
    is_addon: bool = False,
) -> Dict[str, Any]:
    dims_norm = [normalize_dim(d) for d in unit_dims_raw if clean_str(d)]
    return {
        "sheet": sheet,
        "category1": clean_str(category1),
        "category2": clean_str(category2),
        "service": clean_str(service),
        "unit_price": unit_price if unit_price is None else float(unit_price),
        "unit_dims": dims_norm,
        "unit_text": clean_str(unit_text),
        "is_addon": bool(is_addon),
        "priced": bool(priced and unit_price is not None),
        "notes": clean_str(notes),
    }


def build_catalog(xlsx_path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    items: List[Dict[str, Any]] = []

    # 动物饲养相关: A=服务项目, B=价格, C=服务说明
    ws = wb["动物饲养相关"]
    for r in range(2, ws.max_row + 1):
        service = clean_str(ws.cell(r, 1).value)
        price = ws.cell(r, 2).value
        desc = clean_str(ws.cell(r, 3).value)
        if not service:
            continue
        parsed = parse_price_cell(price)
        if parsed["priced"]:
            items.append(make_item(
                "动物饲养相关", service,
                parsed["unit_price"], parsed["dims"], parsed["unit_text"],
                notes=desc, priced=True, is_addon=parsed["is_addon"]
            ))
        else:
            items.append(make_item(
                "动物饲养相关", service,
                None, [], "",
                notes=f"{clean_str(price)} {desc}".strip(),
                priced=False
            ))

    # 造模: A=分类, B=模型名称, C=价格(通常为数值，默认按 只)
    ws = wb["造模"]
    current_cat = ""
    for r in range(2, ws.max_row + 1):
        cat = clean_str(ws.cell(r, 1).value)
        if cat:
            current_cat = cat
        name = clean_str(ws.cell(r, 2).value)
        price = ws.cell(r, 3).value
        if not name:
            continue

        if isinstance(price, (int, float)) and float(price) > 0:
            items.append(make_item(
                "造模", name, float(price),
                ["只"], "元/只",
                category1=current_cat, priced=True
            ))
        else:
            parsed = parse_price_cell(price)
            if parsed["priced"]:
                items.append(make_item(
                    "造模", name, parsed["unit_price"],
                    parsed["dims"] or ["只"],
                    parsed["unit_text"] or "元/只",
                    category1=current_cat, priced=True
                ))
            else:
                items.append(make_item(
                    "造模", name, None, [], "",
                    category1=current_cat, notes=clean_str(price), priced=False
                ))

    # 行为学实验: A=服务项目, B=价格(数值), C=说明, D=场地费用(数值/小时), E=备注
    ws = wb["行为学实验"]
    for r in range(2, ws.max_row + 1):
        service = clean_str(ws.cell(r, 1).value)
        if not service:
            continue
        price = ws.cell(r, 2).value
        desc = clean_str(ws.cell(r, 3).value)
        venue = ws.cell(r, 4).value
        remark = clean_str(ws.cell(r, 5).value)

        if isinstance(price, (int, float)) and float(price) > 0:
            items.append(make_item(
                "行为学实验", service, float(price),
                ["只"], "元/只", notes="; ".join([desc, remark]).strip("; "), priced=True
            ))
        else:
            parsed = parse_price_cell(price)
            items.append(make_item(
                "行为学实验", service,
                parsed["unit_price"], parsed["dims"] or ["只"],
                parsed["unit_text"] or "元/只",
                notes="; ".join([desc, remark]).strip("; "),
                priced=parsed["priced"]
            ))

        if isinstance(venue, (int, float)) and float(venue) > 0:
            items.append(make_item(
                "行为学实验", f"{service}（场地费）", float(venue),
                ["小时"], "元/小时", notes=remark, priced=True
            ))

    # 病理: A=I分类, B=II分类, C=服务项目, D=单位, E=单价(数值)
    ws = wb["病理"]
    for r in range(2, ws.max_row + 1):
        c1 = clean_str(ws.cell(r, 1).value)
        c2 = clean_str(ws.cell(r, 2).value)
        service = clean_str(ws.cell(r, 3).value)
        unit = ws.cell(r, 4).value
        price = ws.cell(r, 5).value
        if not service:
            continue
        u = parse_unit_text(unit)
        if isinstance(price, (int, float)) and float(price) > 0:
            items.append(make_item(
                "病理", service, float(price),
                u["dims"], u["unit_text"],
                category1=c1, category2=c2, priced=True
            ))
        else:
            parsed = parse_price_cell(price)
            items.append(make_item(
                "病理", service, parsed["unit_price"],
                parsed["dims"] or u["dims"],
                parsed["unit_text"] or u["unit_text"],
                category1=c1, category2=c2,
                priced=parsed["priced"]
            ))

    # 荧光定量PCR: A=服务项目, B=实验说明, C=价格(带单位), D=周期, E=备注
    ws = wb["荧光定量PCR"]
    for r in range(2, ws.max_row + 1):
        service = clean_str(ws.cell(r, 1).value)
        if not service:
            continue
        exp = clean_str(ws.cell(r, 2).value)
        price = ws.cell(r, 3).value
        cycle = clean_str(ws.cell(r, 4).value)
        remark = clean_str(ws.cell(r, 5).value)

        parsed = parse_price_cell(price)
        items.append(make_item(
            "荧光定量PCR", service, parsed["unit_price"],
            parsed["dims"], parsed["unit_text"],
            notes="; ".join([exp, cycle, remark]).strip("; "),
            priced=parsed["priced"]
        ))

    # WB: A=服务项目, B=价格(带单位), C=服务说明
    ws = wb["WB"]
    for r in range(2, ws.max_row + 1):
        service = clean_str(ws.cell(r, 1).value)
        if not service:
            continue
        price = ws.cell(r, 2).value
        desc = clean_str(ws.cell(r, 3).value)
        parsed = parse_price_cell(price)
        items.append(make_item(
            "WB", service, parsed["unit_price"],
            parsed["dims"], parsed["unit_text"],
            notes=desc, priced=parsed["priced"]
        ))

    # 组学: A=服务项目, B=价格(带单位), C=周期, D=备注
    ws = wb["组学"]
    for r in range(2, ws.max_row + 1):
        c1 = clean_str(ws.cell(r, 1).value)
        service = clean_str(ws.cell(r, 2).value)
        if not service:
            continue
        price = ws.cell(r, 3).value
        cycle = clean_str(ws.cell(r, 4).value)
        parsed = parse_price_cell(price)
        items.append(make_item(
            "组学", service, parsed["unit_price"],
            parsed["dims"], parsed["unit_text"],
            notes=cycle,
            category1=c1,
            priced=parsed["priced"]
        ))

    # 输出最终 catalog
    return {
        "currency": "CNY",
        "source": Path(xlsx_path).name,
        "items": items,
    }


def catalog_to_prompt_text(catalog: Dict[str, Any]) -> str:
    """
    生成更适合 LLM 阅读/匹配的文本目录。
    """
    lines: List[str] = []
    lines.append(f"价格表目录（来源={catalog.get('source','')}，币种={catalog.get('currency','')}）")
    lines.append("字段含义：服务名 | 单价 | 单位维度 | 单位文本 | sheet | 备注")
    for it in catalog["items"]:
        unit_price = it["unit_price"]
        price_str = "NA" if unit_price is None else str(unit_price)
        dims_str = "×".join(it["unit_dims"]) if it["unit_dims"] else "NA"
        note = it["notes"][:60] if it.get("notes") else ""
        lines.append(
            f"{it['service']} | {price_str} | {dims_str} | {it['unit_text'] or 'NA'} | {it['sheet']} | {note}"
        )
    return "\n".join(lines)


if __name__ == "__main__":
    catalog = build_catalog(XLSX_PATH)

    Path(OUT_JSON).write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    text = catalog_to_prompt_text(catalog)
    Path(OUT_TXT).write_text(text, encoding="utf-8")

    print(f"OK: wrote {OUT_JSON} and {OUT_TXT}. items={len(catalog['items'])}")
