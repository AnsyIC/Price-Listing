from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json
from io import BytesIO
import math
import base64

def generate_excel(pricing_data):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Breakdown"

    # Define headers
    headers = ["服务项目", "单价", "数量", "备注", "总费用"]
    ws.append(headers)

    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Parse pricing report and add data rows
    items = []
    if 'sections' in pricing_data:
        for section in pricing_data['sections']:
            if 'items' in section:
                items.extend(section['items'])
    elif 'items' in pricing_data:
        items = pricing_data['items']
    else:
        items = [pricing_data]

    row_num = 2
    for item in items:
        service = item.get('service', item.get('name', 'N/A'))
        unit_price = item.get('unitPrice', item.get('unitPrice', 0))
        
        # Handle quantityFactor
        quantity = math.prod(item.get('quantityFactors', {"次": 1}).values())
        need_attention = "需要人工检查" if item.get('isOutsourced') else ""

        # Add data to row
        ws.cell(row=row_num, column=1, value=service)
        ws.cell(row=row_num, column=2, value=unit_price)
        ws.cell(row=row_num, column=3, value=quantity)
        ws.cell(row=row_num, column=4, value=need_attention)
        
        # Add formula for Total Cost
        # Note: If quantity is string, this formula will error in Excel. 
        # But we keep the logic from n8n.
        formula = f"=B{row_num}*C{row_num}"
        ws.cell(row=row_num, column=5, value=formula)
        
        # Apply borders and number formatting
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col in [2, 5]:  # Unit Price and Total Cost columns
                cell.number_format = '$#,##0.00'
            if col == 3:  # Quantity column
                cell.alignment = Alignment(horizontal='center')
        
        row_num += 1

    # Add Grand Total row
    ws.cell(row=row_num, column=4, value="Grand Total:")
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')

    # Add SUM formula for grand total
    sum_formula = f"=SUM(E2:E{row_num-1})"
    ws.cell(row=row_num, column=5, value=sum_formula)
    ws.cell(row=row_num, column=5).font = Font(bold=True)
    ws.cell(row=row_num, column=5).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=5).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Apply borders to grand total row
    for col in range(4, 6):
        ws.cell(row=row_num, column=col).border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    # Save workbook to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Convert to base64 for n8n binary data
    excel_data = buffer.getvalue()
    excel_base64 = base64.b64encode(excel_data).decode('utf-8')

    return {
        'filename': 'pricing_breakdown.xlsx',
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'data': excel_base64
    }
